from decimal import Decimal
from io import StringIO

import openpyxl
import pytest
from django.core.management import call_command
from django.core.management.base import CommandError

from budgetis.accounting.management.commands.import_mch2_functional_classification import DEFAULT_EXCEL_PATH
from budgetis.accounting.models import Account
from budgetis.accounting.models import AccountCodeMapping
from budgetis.accounting.models import AccountGroup
from budgetis.accounting.models import GroupResponsibility
from budgetis.accounting.tests.factories import AccountFactory
from budgetis.accounting.tests.factories import AccountGroupFactory
from budgetis.accounting.tests.factories import GroupResponsibilityFactory
from budgetis.common.models import ChartScheme
from budgetis.finance.models import AvailableYear
from budgetis.users.tests.factories import UserFactory


pytestmark = pytest.mark.django_db


def _write_excel(tmp_path, rows):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["No", "Compte", "Resp BUD"])
    for row in rows:
        sheet.append(row)
    path = tmp_path / "responsibles.xlsx"
    workbook.save(path)
    return path


def _write_classification_excel(tmp_path, rows):
    """`rows` are (n1, n2, n3, n4, label) tuples, as they appear in the source sheet."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Classification fonctionnelle"
    sheet.append(["Fonction : Niveau", None, None, None, "Désignation"])
    sheet.append(["N1", "N2", "N3", "N4", None])
    sheet.append(["CLASSIFICATION FONCTIONNELLE", None, None, None, None])
    for row in rows:
        sheet.append(row)
    path = tmp_path / "classification.xlsx"
    workbook.save(path)
    return path


def _run(excel_path, year, *, dry_run=False):
    out, err = StringIO(), StringIO()
    call_command(
        "import_group_responsibilities",
        str(excel_path),
        year=year,
        dry_run=dry_run,
        stdout=out,
        stderr=err,
    )
    return out.getvalue(), err.getvalue()


class TestImportGroupResponsibilities:
    def test_raises_when_file_missing(self, tmp_path):
        with pytest.raises(CommandError):
            call_command("import_group_responsibilities", str(tmp_path / "missing.xlsx"), year=2026)

    def test_creates_responsibility_when_none_existed(self, tmp_path):
        group = AccountGroupFactory(code="720")
        user = UserFactory(trigram="ADA")
        excel_path = _write_excel(tmp_path, [["720.351", "Aide sociale", "ADA"]])

        _run(excel_path, 2026)

        responsibility = GroupResponsibility.objects.get(group=group, year=2026)
        assert responsibility.responsible == user

    def test_dry_run_does_not_persist(self, tmp_path):
        AccountGroupFactory(code="720")
        UserFactory(trigram="ADA")
        excel_path = _write_excel(tmp_path, [["720.351", "Aide sociale", "ADA"]])

        _run(excel_path, 2026, dry_run=True)

        assert not GroupResponsibility.objects.exists()

    def test_unchanged_when_already_correct(self, tmp_path):
        user = UserFactory(trigram="ADA")
        group = AccountGroupFactory(code="720")
        GroupResponsibilityFactory(group=group, year=2026, responsible=user)
        excel_path = _write_excel(tmp_path, [["720.351", "Aide sociale", "ADA"]])

        out, _ = _run(excel_path, 2026)

        assert "0 updated, 1 unchanged, 0 skipped" in out

    def test_updates_existing_responsibility(self, tmp_path):
        old_user = UserFactory(trigram="PCO")
        new_user = UserFactory(trigram="ADA")
        group = AccountGroupFactory(code="720")
        GroupResponsibilityFactory(group=group, year=2026, responsible=old_user)
        excel_path = _write_excel(tmp_path, [["720.351", "Aide sociale", "ADA"]])

        _run(excel_path, 2026)

        responsibility = GroupResponsibility.objects.get(group=group, year=2026)
        assert responsibility.responsible == new_user

    def test_skips_unknown_account_group(self, tmp_path):
        UserFactory(trigram="ADA")
        excel_path = _write_excel(tmp_path, [["999.351", "Inconnu", "ADA"]])

        out, err = _run(excel_path, 2026)

        assert not GroupResponsibility.objects.exists()
        assert "0 updated, 0 unchanged, 1 skipped" in out
        assert "999" in err

    def test_skips_unknown_trigram(self, tmp_path):
        AccountGroupFactory(code="720")
        excel_path = _write_excel(tmp_path, [["720.351", "Aide sociale", "XXX"]])

        out, err = _run(excel_path, 2026)

        assert not GroupResponsibility.objects.exists()
        assert "0 updated, 0 unchanged, 1 skipped" in out
        assert "XXX" in err

    def test_ignores_row_with_blank_trigram(self, tmp_path):
        AccountGroupFactory(code="720")
        excel_path = _write_excel(tmp_path, [["720.351", "Aide sociale", None]])

        out, err = _run(excel_path, 2026)

        assert not GroupResponsibility.objects.exists()
        assert "0 updated, 0 unchanged, 0 skipped" in out
        assert err == ""

    def test_skips_conflicting_trigrams_for_same_function(self, tmp_path):
        AccountGroupFactory(code="720")
        UserFactory(trigram="ADA")
        UserFactory(trigram="PCO")
        excel_path = _write_excel(
            tmp_path,
            [
                ["720.351", "Aide sociale", "ADA"],
                ["720.352", "Autre", "PCO"],
            ],
        )

        out, err = _run(excel_path, 2026)

        assert not GroupResponsibility.objects.exists()
        assert "0 updated, 0 unchanged, 0 skipped" in out
        assert "conflicting trigrams" in err


def _run_classification(excel_path, *, dry_run=False):
    out, err = StringIO(), StringIO()
    call_command(
        "import_mch2_functional_classification",
        str(excel_path),
        dry_run=dry_run,
        stdout=out,
        stderr=err,
    )
    return out.getvalue(), err.getvalue()


class TestImportMch2FunctionalClassification:
    def test_raises_when_file_missing(self, tmp_path):
        with pytest.raises(CommandError):
            call_command("import_mch2_functional_classification", str(tmp_path / "missing.xlsx"))

    def test_raises_readable_error_when_sheet_missing(self, tmp_path):
        workbook = openpyxl.Workbook()
        workbook.active.title = "Fonctionnement"
        path = tmp_path / "wrong-file.xlsx"
        workbook.save(path)

        with pytest.raises(CommandError, match="Available sheets: Fonctionnement"):
            call_command("import_mch2_functional_classification", str(path))

    def test_builds_four_level_hierarchy(self, tmp_path):
        excel_path = _write_classification_excel(
            tmp_path,
            [
                (0, None, None, None, "Administration générale"),
                (None, "01", None, None, "Législatif et exécutif"),
                (None, None, "011", None, "Législatif"),
                (None, None, None, "0110", "Législatif"),
            ],
        )

        out, _ = _run_classification(excel_path)

        assert "4 created, 0 updated, 0 unchanged" in out
        n1 = AccountGroup.objects.get(scheme=ChartScheme.MCH2, level=1, code="0")
        n2 = AccountGroup.objects.get(scheme=ChartScheme.MCH2, level=2, code="01")
        n3 = AccountGroup.objects.get(scheme=ChartScheme.MCH2, level=3, code="011")
        n4 = AccountGroup.objects.get(scheme=ChartScheme.MCH2, level=4, code="0110")
        assert n2.parent == n1
        assert n3.parent == n2
        assert n4.parent == n3
        assert n4.label == "Législatif"

    def test_dry_run_does_not_persist(self, tmp_path):
        excel_path = _write_classification_excel(tmp_path, [(0, None, None, None, "Administration générale")])

        _run_classification(excel_path, dry_run=True)

        assert not AccountGroup.objects.filter(scheme=ChartScheme.MCH2).exists()

    def test_rerun_is_unchanged(self, tmp_path):
        excel_path = _write_classification_excel(tmp_path, [(0, None, None, None, "Administration générale")])
        _run_classification(excel_path)

        out, _ = _run_classification(excel_path)

        assert "0 created, 0 updated, 1 unchanged" in out

    def test_rerun_updates_changed_label(self, tmp_path):
        excel_path = _write_classification_excel(tmp_path, [(0, None, None, None, "Administration générale")])
        _run_classification(excel_path)
        renamed_path = _write_classification_excel(tmp_path, [(0, None, None, None, "Administration générale bis")])

        out, _ = _run_classification(renamed_path)

        assert "0 created, 1 updated, 0 unchanged" in out
        group = AccountGroup.objects.get(scheme=ChartScheme.MCH2, level=1, code="0")
        assert group.label == "Administration générale bis"

    def test_does_not_touch_mch1_groups(self, tmp_path):
        AccountGroupFactory(code="0", scheme=ChartScheme.MCH1, level=1)
        excel_path = _write_classification_excel(tmp_path, [(0, None, None, None, "Administration générale")])

        _run_classification(excel_path)

        assert AccountGroup.objects.filter(scheme=ChartScheme.MCH1).count() == 1
        assert AccountGroup.objects.filter(scheme=ChartScheme.MCH2).count() == 1

    def test_corrects_known_source_typo(self, tmp_path):
        excel_path = _write_classification_excel(
            tmp_path,
            [
                (None, "53", None, None, "Vieillesse et survivants"),
                (None, None, "352", None, "Prestations complémentaires AVS"),
                (None, None, None, "5320", "Prestations complémentaires AVS"),
            ],
        )

        _run_classification(excel_path)

        n2 = AccountGroup.objects.get(scheme=ChartScheme.MCH2, level=2, code="53")
        n3 = AccountGroup.objects.get(scheme=ChartScheme.MCH2, level=3, code="532")
        assert n3.parent == n2
        assert not AccountGroup.objects.filter(scheme=ChartScheme.MCH2, level=3, code="352").exists()

    def test_official_reference_file_has_documented_shape(self):
        """
        Locks in the canton reference file's node counts (10/69/159/180, see
        docs/mch2-migration.md) so a future edition of the file surfaces here
        instead of silently producing a different hierarchy shape.
        """
        if not DEFAULT_EXCEL_PATH.exists():
            pytest.skip("Reference file not checked out")

        _run_classification(DEFAULT_EXCEL_PATH)

        counts = {
            level: AccountGroup.objects.filter(scheme=ChartScheme.MCH2, level=level).count() for level in (1, 2, 3, 4)
        }
        assert counts == {1: 10, 2: 69, 3: 159, 4: 180}


def _write_mapping_excel(tmp_path, rows):
    """`rows` are (mch1_function, mch1_nature, mch1_sub_account, mch2_function, mch2_nature, mch2_sub_account)."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Fonctionnement"
    sheet.append(["ADMIN1", "F", "NAT2", "ADMIN1_N", "NAT1_N", "NAT2_N"])
    for row in rows:
        sheet.append(row)
    path = tmp_path / "conversion.xlsx"
    workbook.save(path)
    return path


def _run_mapping(excel_path, *, dry_run=False):
    out, err = StringIO(), StringIO()
    call_command("import_account_code_mapping", str(excel_path), dry_run=dry_run, stdout=out, stderr=err)
    return out.getvalue(), err.getvalue()


class TestImportAccountCodeMapping:
    def test_raises_when_file_missing(self, tmp_path):
        with pytest.raises(CommandError):
            call_command("import_account_code_mapping", str(tmp_path / "missing.xlsx"))

    def test_raises_readable_error_when_sheet_missing(self, tmp_path):
        workbook = openpyxl.Workbook()
        workbook.active.title = "Classification fonctionnelle"
        path = tmp_path / "wrong-file.xlsx"
        workbook.save(path)

        with pytest.raises(CommandError, match="Available sheets: Classification fonctionnelle"):
            call_command("import_account_code_mapping", str(path))

    def test_prompts_and_uses_chosen_sheet_when_interactive(self, tmp_path, monkeypatch):
        excel_path = _write_mapping_excel(tmp_path, [(100, 301, 0, 1100, 3010, 0)])
        workbook = openpyxl.load_workbook(excel_path)
        workbook["Fonctionnement"].title = "Fonctionnement (nouveau)"
        workbook.save(excel_path)
        monkeypatch.setattr("sys.stdin.isatty", lambda: True)
        monkeypatch.setattr("builtins.input", lambda prompt: "1")

        out, _ = _run_mapping(excel_path)

        assert "1 created, 0 unchanged" in out
        assert AccountCodeMapping.objects.exists()

    def test_creates_mapping_from_both_sides_filled(self, tmp_path):
        excel_path = _write_mapping_excel(tmp_path, [(100, 301, 0, 1100, 3010, 0)])

        out, _ = _run_mapping(excel_path)

        assert "1 created, 0 unchanged" in out
        mapping = AccountCodeMapping.objects.get()
        assert (mapping.mch1_function, mapping.mch1_nature, mapping.mch1_sub_account) == ("100", "301", "")
        assert (mapping.mch2_function, mapping.mch2_nature, mapping.mch2_sub_account) == ("01100", "3010", "")

    def test_zero_pads_mch2_codes_dropped_by_pandas_int_coercion(self, tmp_path):
        # Pandas reads a numeric-looking column as int, silently dropping the
        # leading zero that makes up the commune-specific digit (e.g. "01100").
        excel_path = _write_mapping_excel(tmp_path, [(100, 301, 0, 1100, 3010, 1)])

        _run_mapping(excel_path)

        mapping = AccountCodeMapping.objects.get()
        assert mapping.mch2_function == "01100"
        assert mapping.mch2_sub_account == "01"

    def test_preserves_a_real_mch1_sub_account(self, tmp_path):
        excel_path = _write_mapping_excel(tmp_path, [(110, 365, 1, 59200, 3632, 0)])

        _run_mapping(excel_path)

        mapping = AccountCodeMapping.objects.get()
        assert mapping.mch1_sub_account == "1"

    def test_skips_row_missing_mch1_side(self, tmp_path):
        excel_path = _write_mapping_excel(tmp_path, [(None, None, None, 1100, 3052, 0)])

        _run_mapping(excel_path)

        assert not AccountCodeMapping.objects.exists()

    def test_skips_fully_blank_row(self, tmp_path):
        excel_path = _write_mapping_excel(tmp_path, [(None, None, None, None, None, None)])

        out, _ = _run_mapping(excel_path)

        assert "0 created, 0 unchanged" in out

    def test_dry_run_does_not_persist(self, tmp_path):
        excel_path = _write_mapping_excel(tmp_path, [(100, 301, 0, 1100, 3010, 0)])

        _run_mapping(excel_path, dry_run=True)

        assert not AccountCodeMapping.objects.exists()

    def test_rerun_is_idempotent(self, tmp_path):
        excel_path = _write_mapping_excel(tmp_path, [(100, 301, 0, 1100, 3010, 0)])
        _run_mapping(excel_path)

        out, _ = _run_mapping(excel_path)

        assert "0 created, 1 unchanged" in out

    def test_split_source_produces_two_mapping_rows(self, tmp_path):
        excel_path = _write_mapping_excel(
            tmp_path,
            [
                (100, 306, 0, 1100, 3049, 0),
                (100, 306, 0, 1100, 3099, 0),
            ],
        )

        out, _ = _run_mapping(excel_path)

        assert "2 created, 0 unchanged" in out
        assert AccountCodeMapping.objects.filter(mch1_function="100", mch1_nature="306").count() == 2  # noqa: PLR2004


def _write_accounts_excel(tmp_path, rows):
    """`rows` are (mch2_function, mch2_nature, mch2_sub_account, label)."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Fonctionnement"
    sheet.append(["ADMIN1_N", "NAT1_N", "NAT2_N", "LIBELLÉ_N"])
    for row in rows:
        sheet.append(row)
    path = tmp_path / "accounts.xlsx"
    workbook.save(path)
    return path


def _run_accounts(excel_path, year, *, actuals=False, dry_run=False):
    out, err = StringIO(), StringIO()
    call_command(
        "import_mch2_accounts",
        str(excel_path),
        year=year,
        actuals=actuals,
        dry_run=dry_run,
        stdout=out,
        stderr=err,
    )
    return out.getvalue(), err.getvalue()


class TestImportMch2Accounts:
    def test_raises_when_file_missing(self, tmp_path):
        with pytest.raises(CommandError):
            call_command("import_mch2_accounts", str(tmp_path / "missing.xlsx"), year=2027)

    def test_creates_zero_value_budget_accounts(self, tmp_path):
        excel_path = _write_accounts_excel(tmp_path, [(1100, 3010, 0, "Salaire")])

        out, _ = _run_accounts(excel_path, 2027)

        assert "1 created, 0 already existed" in out
        account = Account.objects.get()
        assert (account.function, account.nature, account.sub_account) == ("01100", "3010", "")
        assert account.label == "Salaire"
        assert account.scheme == ChartScheme.MCH2
        assert account.is_budget is True
        assert account.charges == Decimal("0.00")
        assert account.revenues == Decimal("0.00")

    def test_registers_available_year_as_mch2(self, tmp_path):
        excel_path = _write_accounts_excel(tmp_path, [(1100, 3010, 0, "Salaire")])

        _run_accounts(excel_path, 2027)

        available_year = AvailableYear.objects.get(year=2027, type=AvailableYear.YearType.BUDGET)
        assert available_year.scheme == ChartScheme.MCH2

    def test_actuals_flag_creates_actuals_instead_of_budget(self, tmp_path):
        excel_path = _write_accounts_excel(tmp_path, [(1100, 3010, 0, "Salaire")])

        _run_accounts(excel_path, 2027, actuals=True)

        account = Account.objects.get()
        assert account.is_budget is False
        assert AvailableYear.objects.filter(year=2027, type=AvailableYear.YearType.ACTUAL).exists()

    def test_derives_expected_type_from_nature_prefix(self, tmp_path):
        excel_path = _write_accounts_excel(
            tmp_path,
            [
                (1100, 3010, 0, "Charge"),
                (1100, 4010, 0, "Revenu"),
            ],
        )

        _run_accounts(excel_path, 2027)

        charge = Account.objects.get(nature="3010")
        revenue = Account.objects.get(nature="4010")
        assert charge.expected_type == Account.ExpectedType.CHARGE
        assert revenue.expected_type == Account.ExpectedType.REVENUE

    def test_deduplicates_repeated_mch2_target_from_a_merge(self, tmp_path):
        excel_path = _write_accounts_excel(
            tmp_path,
            [
                (96900, 3420, 0, "Fusion"),
                (96900, 3420, 0, "Fusion"),
            ],
        )

        out, _ = _run_accounts(excel_path, 2027)

        assert "1 created, 0 already existed" in out
        assert Account.objects.count() == 1

    def test_dry_run_does_not_persist(self, tmp_path):
        excel_path = _write_accounts_excel(tmp_path, [(1100, 3010, 0, "Salaire")])

        _run_accounts(excel_path, 2027, dry_run=True)

        assert not Account.objects.exists()
        assert not AvailableYear.objects.exists()

    def test_never_overwrites_an_existing_account(self, tmp_path):
        AccountFactory(
            scheme=ChartScheme.MCH2,
            year=2027,
            function="01100",
            nature="3010",
            sub_account="",
            is_budget=True,
            label="Real imported data",
            charges=Decimal("5000.00"),
        )
        excel_path = _write_accounts_excel(tmp_path, [(1100, 3010, 0, "Salaire")])

        out, _ = _run_accounts(excel_path, 2027)

        assert "0 created, 1 already existed" in out
        account = Account.objects.get()
        assert account.label == "Real imported data"
        assert account.charges == Decimal("5000.00")
